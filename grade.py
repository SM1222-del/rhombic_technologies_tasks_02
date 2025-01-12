from openpyxl import Workbook, load_workbook
import zipfile
import hashlib

# Student management system using CRUD operations
"""
add
update
delete
view
exit
"""

class Student:
    def __init__(self, username, password, name, student_id):
        self.username = username
        self.password = password
        self.name = name
        self.student_id = student_id

    # Create a new Excel file with headers if it doesn't exist
    @staticmethod
    def create_excel_file(file_name):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Grades"
        sheet.append(["Student Name", "Student ID", "Subject", "Grade"])
        wb.save(file_name)
        print(f"New file '{file_name}' created.")

    # View grades for the student (students only have view access)
    @staticmethod
    def view_grades(file_name, student_id):
        try:
            wb = load_workbook(file_name)
            sheet = wb["Grades"]
        except FileNotFoundError:
            print(f"File '{file_name}' not found.")
            return

        # Search and display the grades for this student
        print(f"Grades for Student ID {student_id}:")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == student_id:
                print(f"Subject: {row[2]}, Grade: {row[3]}")

    # View attendance (for future implementation)
    @staticmethod
    def view_attendance(student_id):
        print(f"Viewing attendance for Student ID {student_id}... (functionality not implemented)")

    # View portfolio (for future implementation)
    @staticmethod
    def view_portfolio(student_id):
        print(f"Viewing portfolio for Student ID {student_id}... (functionality not implemented)")


class Teacher:
    def __init__(self, username, password):
        self.username = username
        self.password = password

    # Add a student to the users list
    def add_student(self, users, name, student_id, student_username, student_password):
        student = Student(student_username, student_password, name, student_id)
        users.append(student)
        print("Student added successfully.")

    # Add or update grade for a student
    def add_or_update_grade(self, file_name, student_name, student_id, subject, grade):
        try:
            wb = load_workbook(file_name)
            sheet = wb["Grades"]
        except FileNotFoundError:
            print(f"File '{file_name}' not found.")
            return
        except zipfile.BadZipFile:
            print(f"Error: '{file_name}' is not a valid Excel file or is corrupted.")
            return
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            return

        # Check if the grade for this student and subject already exists
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[1].value == student_id and row[2].value == subject:
                # Update the grade if the student and subject match
                row[3].value = grade
                print(f"Grade for {student_name} in {subject} updated to {grade}.")
                wb.save(file_name)
                return

        # If grade doesn't exist, add a new row
        sheet.append([student_name, student_id, subject, grade])
        print(f"Grade for {student_name} in {subject} added with grade {grade}.")
        wb.save(file_name)

    # Function to delete grade in the Excel file
    def delete_grade(self, file_name, student_id, subject):
        try:
            wb = load_workbook(file_name)
            sheet = wb["Grades"]
        except FileNotFoundError:
            print(f"File '{file_name}' not found.")
            return

        # Search for the grade entry and remove it
        for row in sheet.iter_rows(min_row=2):
            if row[1].value == student_id and row[2].value == subject:
                sheet.delete_rows(row[0].row)  # Delete the row with matching student_id and subject
                print(f"Grade for student ID {student_id} in {subject} deleted.")
                wb.save(file_name)
                return
        print(f"Grade for student ID {student_id} in {subject} not found.")

    # Mark attendance for a student
    def mark_attendance(self, users, student_id, date, status):
        for user in users:
            if isinstance(user, Student) and user.student_id == student_id:
                print(f"Attendance for student ID {student_id} on {date}: {status}")
                return
        print("Student not found.")

    # Add a portfolio entry for a student
    def add_portfolio_entry(self, users, student_id, entry_type, entry):
        for user in users:
            if isinstance(user, Student) and user.student_id == student_id:
                print(f"Portfolio entry for student ID {student_id}: {entry_type} - {entry}")
                return
        print("Student not found.")

    # List all students
    def list_all_students(self, users):
        print("List of Students:")
        for user in users:
            if isinstance(user, Student):
                print(f"ID: {user.student_id}, Name: {user.name}")

    # Get information of a specific student
    def get_student_information(self, users, student_id):
        for user in users:
            if isinstance(user, Student) and user.student_id == student_id:
                print(f"Name: {user.name}")
                print(f"ID: {user.student_id}")
                return
        print("Student not found.")

    # Calculate the average grade for all students
    def calculate_average(self, file_name):
        try:
            wb = load_workbook(file_name)
            sheet = wb["Grades"]
        except FileNotFoundError:
            print(f"File '{file_name}' not found.")
            return

        grades = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            grade = row[3]  # Get the grade column (index 3)
            if grade is not None:
                try:
                    # Check if the grade is a string with commas (indicating multiple grades)
                    if isinstance(grade, str) and ',' in grade:
                        # Split by comma and convert to float each part
                        grade_values = grade.split(',')
                        for val in grade_values:
                            grades.append(float(val.strip()))  # Convert each value to float
                    else:
                        grades.append(float(grade))  # Convert grade to float
                except ValueError:
                    print(f"Invalid grade value: '{grade}' in row {row}")
                    continue  # Skip invalid grades

        if grades:
            average = sum(grades) / len(grades)
            print(f"The average grade is: {average:.2f}")
        else:
            print("No valid grades available to calculate the average.")

# Hashing passwords for security
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# Authentication function to log in
def authenticate(username, password, users):
    for user in users:
        if user.username == username and user.password == hash_password(password):
            return user
    return None


def main():
    users = [Teacher("Sidra", hash_password("123"))]  # default teacher name & password
    file_name = "grades.xlsx"

# Create the Excel file if it doesn't exist
    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        Student.create_excel_file(file_name)

    while True:
        print("\nStudent Grade and Academic Performance Tracking System")
        print("1. Student Login")
        print("2. Teacher Login")
        print("3. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            username = input("Enter username: ")
            password = input("Enter password: ")
            user = authenticate(username, password, users)
            if isinstance(user, Student):
                print("Welcome, Student!")
                while True:
                    print("\nStudent Menu")
                    print("1. View Grades")
                    print("2. View Attendance")
                    print("3. View Portfolio")
                    print("4. Exit")
                    student_choice = input("Enter your choice: ")
                    if student_choice == "1":
                        Student.view_grades(file_name, user.student_id)
                    elif student_choice == "2":
                        Student.view_attendance(user.student_id)
                    elif student_choice == "3":
                        Student.view_portfolio(user.student_id)
                    elif student_choice == "4":
                        break
                    else:
                        print("Invalid choice. Please try again.")

            else:
                print("Invalid username or password.")

        elif choice == "2":
            username = input("Enter username: ")
            password = input("Enter password: ")
            user = authenticate(username, password, users)
            if isinstance(user, Teacher):
                print("Welcome, Teacher!")
                while True:
                    print("\nTeacher Menu")
                    print("1. Add Student")
                    print("2. Add/Update Student Grade")
                    print("3. Delete Grade")
                    print("4. Mark Attendance")
                    print("5. Add Portfolio Entry")
                    print("6. List All Students")
                    print("7. Get Student Information")
                    print("8. Calculate Average Grade")
                    print("9. Exit")
                    teacher_choice = input("Enter your choice: ")
                    if teacher_choice == "1":
                        name = input("Enter student name: ")
                        student_id = input("Enter student ID: ")
                        student_username = input("Enter student username: ")
                        student_password = input("Enter student password: ")
                        user.add_student(users, name, student_id, student_username, student_password)
                    elif teacher_choice == "2":
                        student_name = input("Enter student name: ")
                        student_id = input("Enter student ID: ")
                        subject = input("Enter subject: ")
                        grade = input("Enter grade: ")
                        user.add_or_update_grade(file_name, student_name, student_id, subject, grade)
                    elif teacher_choice == "3":
                        student_id = input("Enter student ID: ")
                        subject = input("Enter subject: ")
                        user.delete_grade(file_name, student_id, subject)
                    elif teacher_choice == "4":
                        student_id = input("Enter student ID: ")
                        date = input("Enter date (YYYY-MM-DD): ")
                        status = input("Enter attendance status (Present/Absent): ")
                        user.mark_attendance(users, student_id, date, status)
                    elif teacher_choice == "5":
                        student_id = input("Enter student ID: ")
                        entry_type = input("Enter entry type (Essay, Project, Assignment): ")
                        entry = input("Enter portfolio entry: ")
                        user.add_portfolio_entry(users, student_id, entry_type, entry)
                    elif teacher_choice == "6":
                        user.list_all_students(users)
                    elif teacher_choice == "7":
                        student_id = input("Enter student ID: ")
                        user.get_student_information(users, student_id)
                    elif teacher_choice == "8":
                        user.calculate_average(file_name)
                    elif teacher_choice == "9":
                        break
                    else:
                        print("Invalid choice. Please try again.")
            else:
                print("Invalid username or password.")

        elif choice == "3":
            print("Exiting...")
            break

        else:
            print("Invalid choice. Please try again.")


if __name__ == "__main__":
    main()
